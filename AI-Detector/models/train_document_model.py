# train_document_model.py
"""
Train a real-vs-fake document classifier using:
 - text from PDFs and Excel files (TF-IDF)
 - numeric metadata features (page count, file size, formula ratio, etc.)
Saves:
 - document_model_calibrated.pkl
 - tfidf_vectorizer.pkl
 - numeric_scaler.pkl
"""

import os
from pathlib import Path
import joblib
import numpy as np

# Text extraction
import fitz  # PyMuPDF
from openpyxl import load_workbook

# ML
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.svm import LinearSVC, SVC
from sklearn.calibration import CalibratedClassifierCV
from scipy.sparse import hstack, csr_matrix
from sklearn.metrics import classification_report, confusion_matrix

# ---------------- CONFIG ----------------
DATASET_DIR = Path("documents_dataset")
REAL_DIR = DATASET_DIR / "real"
FAKE_DIR = DATASET_DIR / "fake"

MODEL_FILE = "document_model_calibrated.pkl"
VECTORIZER_FILE = "tfidf_vectorizer.pkl"
SCALER_FILE = "numeric_scaler.pkl"

# ---------------- HELPERS ----------------
def extract_text_from_pdf(path: Path):
    try:
        doc = fitz.open(str(path))
        text_blocks = []
        fonts = set()
        for page in doc:
            text_blocks.append(page.get_text())
            try:
                for block in page.get_text("dict")["blocks"]:
                    if "lines" in block:
                        for line in block["lines"]:
                            for span in line["spans"]:
                                fonts.add(span.get("font", ""))
            except Exception:
                pass
        full_text = "\n".join(text_blocks)
        num_pages = doc.page_count
        doc.close()
        return full_text, num_pages, len(fonts)
    except Exception:
        return "", 0, 0

def extract_text_from_excel(path: Path):
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=False)
        texts = []
        total_cells = 0
        formula_cells = 0
        for sheet in wb:
            for row in sheet.iter_rows(values_only=False):
                for cell in row:
                    total_cells += 1
                    val = cell.value
                    if val is None:
                        continue
                    texts.append(str(val))
                    if isinstance(val, str) and val.startswith("="):
                        formula_cells += 1
        wb.close()
        text = "\n".join(texts)
        num_sheets = len(wb.sheetnames)
        formula_ratio = (formula_cells / total_cells) if total_cells > 0 else 0.0
        return text, num_sheets, formula_ratio
    except Exception:
        return "", 0, 0.0

def extract_features(path: Path):
    suffix = path.suffix.lower()
    file_size = path.stat().st_size
    text = ""
    num_pages = 0
    num_fonts = 0
    num_sheets = 0
    formula_ratio = 0.0

    if suffix == ".pdf":
        text, num_pages, num_fonts = extract_text_from_pdf(path)
    elif suffix in [".xlsx", ".xlsm", ".xls"]:
        text, num_sheets, formula_ratio = extract_text_from_excel(path)
    else:
        try:
            text, num_pages, num_fonts = extract_text_from_pdf(path)
        except Exception:
            try:
                with open(path, "r", encoding="utf8", errors="ignore") as f:
                    text = f.read()
            except Exception:
                text = ""

    numeric = [
        file_size / 1024.0,
        num_pages,
        num_fonts,
        num_sheets,
        formula_ratio,
    ]
    return text, numeric

# ---------------- BUILD DATA ----------------
rows = []
for p in REAL_DIR.glob("*"):
    rows.append((p, "real"))
for p in FAKE_DIR.glob("*"):
    rows.append((p, "fake"))

if not rows:
    raise SystemExit("No files found in documents_dataset/real or .../fake.")

paths, labels = zip(*rows)

texts = []
numeric_feats = []
for p in paths:
    t, n = extract_features(p)
    texts.append(t)
    numeric_feats.append(n)

X_num = np.array(numeric_feats, dtype=float)
y = np.array([1 if lab=="fake" else 0 for lab in labels])  # 1=fake,0=real

# ---------------- TEXT VECTORIZATION ----------------
tfidf = TfidfVectorizer(max_features=5000, stop_words="english")
X_text = tfidf.fit_transform(texts)

# ---------------- SCALE NUMERIC ----------------
scaler = StandardScaler()
X_num_scaled = scaler.fit_transform(X_num)
X_num_sparse = csr_matrix(X_num_scaled)

# ---------------- COMBINE ----------------
X = hstack([X_text, X_num_sparse])

# ---------------- SPLIT ----------------
X_train, X_test, y_train, y_test, paths_train, paths_test = train_test_split(
    X, y, paths, test_size=0.2, random_state=42, stratify=y
)

# ---------------- TRAIN BASE LINEAR SVC ----------------
base_svc = LinearSVC(max_iter=20000)
base_svc.fit(X_train, y_train)

# ---------------- CALIBRATE TO GET PROBABILITIES ----------------
calibrated_model = CalibratedClassifierCV(base_svc, method='sigmoid', cv='prefit')
# small calibration subset
X_calib, _, y_calib, _ = train_test_split(X_train, y_train, test_size=0.8, random_state=42)
calibrated_model.fit(X_calib, y_calib)

# ---------------- EVALUATE ----------------
y_pred = calibrated_model.predict(X_test)
y_proba = calibrated_model.predict_proba(X_test)
print("=== Classification Report ===")
print(classification_report(y_test, y_pred, target_names=["real","fake"]))
print("=== Confusion Matrix ===")
print(confusion_matrix(y_test, y_pred))

# ---------------- SAVE ARTIFACTS ----------------
joblib.dump(calibrated_model, MODEL_FILE)
joblib.dump(tfidf, VECTORIZER_FILE)
joblib.dump(scaler, SCALER_FILE)
print(f"Saved calibrated model -> {MODEL_FILE}")
print(f"Saved TF-IDF -> {VECTORIZER_FILE}")
print(f"Saved scaler -> {SCALER_FILE}")

# ---------------- PREDICTION FUNCTION ----------------
def predict_document(path: str):
    """
    Returns: dict {"label": "real"/"fake", "probability": 0.69}
    """
    mdl = joblib.load(MODEL_FILE)
    vec = joblib.load(VECTORIZER_FILE)
    sc = joblib.load(SCALER_FILE)

    text, numeric = extract_features(Path(path))
    X_text_vect = vec.transform([text])
    X_num_scaled = sc.transform([numeric])
    X_in = hstack([X_text_vect, csr_matrix(X_num_scaled)])

    pred_idx = int(mdl.predict(X_in)[0])
    proba = mdl.predict_proba(X_in)[0][pred_idx]

    return {
        "label": "Fake Content" if pred_idx==1 else "Authentic Content",
        "confidence": round(float(proba)*100, 2)
    }
